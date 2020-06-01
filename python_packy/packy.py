import json
import os
from shutil import copyfile
from shutil import rmtree
import openpyxl
import json
import EB_AFIT_core as core


# Генерация конфигурационного файла для запуска упаковки неупакованных коробок
def not_packed(json_data):
    file = open('boxlist.txt', 'w')
    a = []

    a.append(json_data['pallet']['pallet_x'])
    a.append(json_data['pallet']['pallet_y'])
    a.append(json_data['pallet']['pallet_z'])

    file.write(str(a).replace('[', '').replace(']', '') + '\n')

    for i in json_data['not_packed']:
        if not i:
            break
        spc = str(i[0]).replace(' ', '') # Из-за особенностей названий в Си, удаляем пробелы
        file.write(f'{spc} {i[1]}, {i[2]}, {i[3]}, {i[4]}, 1' + '\n')

    file.close()

    return 0


# Создание списка упакованных коробок с разбивкой по паллетам
def boxes(json_data, t):
    pallet_file = open('Robot/Pallet_' + str(t) + '.json', 'w')
    a = []
    b = []

    a.append(json_data['pallet']['pallet_x'])
    a.append(json_data['pallet']['pallet_y'])
    a.append(json_data['pallet']['pallet_z'])

    dict_packed['pallet'] = a

    for i in json_data['boxes']:
        b.append(i)

    dict_packed['item'] = b
    pallet_file.write(str(dict_packed).replace('\'', '"').replace(' ', '') + '\n')
    pallet_file.close()

    return 0


# Создание отчётов Excel для оператора
def excel(t):
    print(json.dumps(templates, sort_keys=False, indent=4))
    for i in templates['boxes']:
        print(i)
    wb = openpyxl.load_workbook(filename = 'xl.xlsx')
    sheet = wb['test']

    sheet.cell(row=2, column=1).value = templates['pallet']['pallet_x']
    sheet.cell(row=2, column=2).value = templates['pallet']['pallet_y']
    sheet.cell(row=2, column=3).value = templates['pallet']['pallet_z']

    i = 1
    q = 1
    for rec in templates['boxes']:
        if not rec:
            i += 1
        for j in rec:
            sheet.cell(row=5 + i, column=q).value = j
            q += 1
        i += 1
        q = 1
    # сохраняем данные по каждой мультипаллете в отдельный xlsx
    name = 'Report/Pallet_'+ str(t) + '.xlsx'
    wb.save(name)
    wb.close()


if __name__ == "__main__":

    for file in os.scandir('Report'):
        if file.name.endswith(".xlsx"):
            os.unlink(file.path)

    for file in os.scandir('Robot'):
        if file.name.endswith(".json"):
            os.unlink(file.path)

    dict_packed = {}

    # Дублируем чтобы не потерять начальную настройку
    copyfile('box.txt', 'boxlist.txt')

    t = 1
    filename = 'boxlist.txt'

    filename_with_path = os.path.abspath(os.path.join(os.path.dirname(__file__), filename))
    core.calc(filename_with_path)

    with open('.~3d') as f:
        templates = json.load(f)
    not_packed(templates)
    boxes(templates, t)
    excel(t)
    t += 1


    while len(templates['not_packed']) > 1 :
        # отправка названия файла в код библиотеки на Си
        with open('.~3d') as f:
            templates = json.load(f)

        filename_with_path = os.path.abspath(os.path.join(os.path.dirname(__file__), filename))
        core.calc(filename_with_path)
        not_packed(templates)
        boxes(templates, t)
        excel(t)
        t += 1

    # Удаление временных файлов, но если немного дописать код, то их можно использовать
    # как источник доп. информации для оператора
    os.remove(".~3d")
    os.remove("boxlist.txt")
    os.remove("boxlist.txt.out") # Простой отчёт, оставлен из оригинального документа

